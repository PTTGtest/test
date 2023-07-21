# coding:utf-8
# @Time     : 2023年06月08日
# @Author   : junxian.geng
# @Email    : junxian.geng@transsion.com
# @description:
import os
import threading

import openpyxl
from openpyxl.reader.excel import load_workbook


# 获取当前手机刷机系统版本号
def get_phone_current_verison():
    cmd = 'adb shell getprop ro.build.display.id'
    result = os.popen(cmd).read().strip()
    return result

# 获取手机当前安卓版本
def get_phone_android_version():
    cmd = 'adb shell getprop ro.build.version.release'
    result = os.popen(cmd).read().strip()
    return result

# 获取手机电池电量
def get_phone_battery_level():
    cmd = 'adb shell dumpsys battery | findstr level'
    result = os.popen(cmd).read().strip()
    return result.split(':')[1].strip()

# 获取当前手机是否是充电状态
def get_phone_battery_status():
    cmd = 'adb shell dumpsys battery | findstr status'
    result = os.popen(cmd).read().strip()
    return result.split(':')[1].strip()

# 获取当前手机连接的wifi名称
def get_phone_wifi_name():
    cmd = 'adb shell dumpsys wifi | findstr SSID'
    result = os.popen(cmd).read().strip()
    return result.split(':')[1].strip()

# 新建表格
def create_excel(path):
    """
    新建表格
    :param path: 表格路径
    :return:
    """
    if not os.path.exists(path):
        wb = openpyxl.Workbook()
        wb.save(path)
    else:
        print('文件已存在')

# 写入内容到表格，并添加线程锁和try-except保证线程安全
lock = threading.Lock()
def write_excel(path, sheet_name, row, col, value):
    """
    写入内容到表格
    :param path: 表格路径
    :param sheet_name: 表格sheet名称
    :param row: 行
    :param col: 列
    :param value: 写入值
    :return:
    """
    # 使用openpyxl写表格导出并添加异常处理和线程锁
    try:
        with lock:
            wb = load_workbook(path)
            ws = wb[sheet_name]
            ws.cell(row=row, column=col, value=value)
            wb.save(path)
            wb.close()
            print('写入成功')
            return True
    except Exception as e:
        print(e)
        return False










